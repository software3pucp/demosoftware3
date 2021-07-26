import random

from django.shortcuts import render
from django.http.response import HttpResponse
from django.http import JsonResponse
from django.core import serializers
from openpyxl import Workbook,load_workbook,styles
import os
from django.conf import settings
from django.db.models import Count
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
# Create your views here.
from demosoftware3.settings import MEDIA_URL
from gestionarEvaluacion.models import RespuestaEvaluacion
from gestionarHorario.models import Horario
from gestionarEspecialidad.models import Especialidad
from gestionarIndicadores.models import Indicador
from gestionarRubrica.models import Rubrica
from gestionarSemestre.models import Semestre
from gestionarEvidencias.models import EvidenciasxHorario
from gestionarNiveles.models import Nivel
from gestionarResultados.models import PlanResultados
from gestionarCurso.models import Curso
from authentication.models import User
from gestionarResultados.models import ResultadoPUCP
from gestionarPlanMedicion.models import PlanMedicionCurso, PlanMedicion
from django.contrib.auth.decorators import login_required
import xlrd
import pandas as pd
@login_required
def evaluar(request,pk):
    media_path = MEDIA_URL
    listaAlumno = reversed(RespuestaEvaluacion.objects.filter(estado=1))
    cursoSeleccionado = Curso.objects.get(pk=pk)
    #REVISAR ESTA PARTE PORQUE FALTA EL ID DE PLAN DE MEDICION
    plan = PlanMedicionCurso.objects.get(estado='1',curso_id=cursoSeleccionado.pk, semestre_id=request.GET['sem'])
    planMedicion = PlanMedicion.objects.get(pk = plan.planMedicion_id)
    #horarios = plan.horario.all()
    horarios = Horario.objects.filter(curso_id=plan.pk)
   # listaHorario= Horario.objects.filter(curso_id=pk) #listaDeHorario asociado a un curso
    especialidadPk = cursoSeleccionado.especialidad_id
    planResultadoPk = PlanResultados.objects.get(especialidad_id=especialidadPk,estado='1')
    resultado = list(ResultadoPUCP.objects.filter(planResultado_id=planResultadoPk))

    if resultado:
        indicadores = plan.indicador.all()
        listaIndicador = list(indicadores)
        #listaIndicador = Indicador.objects.filter(resultado_id=resultado[0].id ,estado=1)
    else:
        listaIndicador = []
    listaHorarios = list(horarios)
    #listaHorarios = list(Horario.objects.filter(curso_id=pk, estado=1))
    # listaDocumentos = EvidenciasxHorario.objects.filter(estado=1)
    listaDocumentos = EvidenciasxHorario.objects.none()
    listaNombres = []
    for i in range(len(listaHorarios)):
        listaEDocumentos = EvidenciasxHorario.objects.filter(horario_id=listaHorarios[i].id, estado=1)
        if listaEDocumentos:
            aux = list(listaEDocumentos)
            for j in range(len(aux)):
                nombArchivo = str(aux[j].archivo)[8:]
                listaNombres.append(nombArchivo)
            # listaEDocumentos.update(archivo=nombArchivo)
            listaDocumentos = listaDocumentos | listaEDocumentos
    cantidad = len(listaNombres)
    listaConjunta = zip(listaNombres, listaDocumentos)
    listaArchivos = set(listaConjunta)
    semestre = Semestre.objects.get(pk=request.GET['sem'])
    print(str(semestre.pk)+ ' '+str(semestre.nombreCodigo))
    context = {
        'media_path': media_path,
        'listaAlumno': listaAlumno,
        'planMedicionCurso': plan,
        'planMedicion': planMedicion,
        'listaIndicador': listaIndicador,
        'cursoSeleccionado': cursoSeleccionado,
        'listaHorario': listaHorarios,
        'listaDocumentos': listaDocumentos,
        'listaNombres': listaNombres,
        'cantidad': cantidad,
        'users': User.objects.all(),
        'listaAux': listaArchivos,
        'semestre':semestre
    }
    return render(request, 'gestionarEvaluacion/baseEvaluacion/base.html',context)

@login_required
def evaluarDocente(request):
    pk = request.POST['cursoButton']
    media_path = MEDIA_URL
    listaAlumno = reversed(RespuestaEvaluacion.objects.filter(estado=1))
    cursoSeleccionado = Curso.objects.get(pk=pk)
    # REVISAR ESTA PARTE PORQUE FALTA EL ID DE PLAN DE MEDICION
    planes = PlanMedicionCurso.objects.filter(estado='1', curso_id=cursoSeleccionado.pk, semestre_id=request.POST['estado'])
    for item in planes:
        if item.planMedicion.estado == '1':
            plan = item
    planMedicion = PlanMedicion.objects.get(pk=plan.planMedicion_id)
    # horarios = plan.horario.all()
    horarios = Horario.objects.filter(curso_id=plan.pk, responsable_id=request.user.pk)
    # listaHorario= Horario.objects.filter(curso_id=pk) #listaDeHorario asociado a un curso
    especialidadPk = cursoSeleccionado.especialidad_id
    planResultadoPk = PlanResultados.objects.get(especialidad_id=especialidadPk, estado='1')
    resultado = list(ResultadoPUCP.objects.filter(planResultado_id=planResultadoPk))

    if resultado:
        indicadores = plan.indicador.all()
        listaIndicador = list(indicadores)
        # listaIndicador = Indicador.objects.filter(resultado_id=resultado[0].id ,estado=1)
    else:
        listaIndicador = []
    listaHorarios = list(horarios)
    # listaHorarios = list(Horario.objects.filter(curso_id=pk, estado=1))
    # listaDocumentos = EvidenciasxHorario.objects.filter(estado=1)
    listaDocumentos = EvidenciasxHorario.objects.none()
    listaNombres = []
    for i in range(len(listaHorarios)):
        listaEDocumentos = EvidenciasxHorario.objects.filter(horario_id=listaHorarios[i].id, estado=1)
        if listaEDocumentos:
            aux = list(listaEDocumentos)
            for j in range(len(aux)):
                nombArchivo = str(aux[j].archivo)[8:]
                listaNombres.append(nombArchivo)
            # listaEDocumentos.update(archivo=nombArchivo)
            listaDocumentos = listaDocumentos | listaEDocumentos
    cantidad = len(listaNombres)
    listaConjunta = zip(listaNombres, listaDocumentos)
    listaArchivos = set(listaConjunta)
    context = {
        'media_path': media_path,
        'listaAlumno': listaAlumno,
        'planMedicionCurso': plan,
        'planMedicion': planMedicion,
        'listaIndicador': listaIndicador,
        'cursoSeleccionado': cursoSeleccionado,
        'listaHorario': listaHorarios,
        'listaDocumentos': listaDocumentos,
        'listaNombres': listaNombres,
        'cantidad': cantidad,
        'users': User.objects.all(),
        'listaAux': listaArchivos,
    }
    print("-------------++-+-+-+-++--+-+-+-+-++-++++++++++++++++++++++++++++++++++++++-------------------")
    print(context)
    return render(request, 'gestionarEvaluacion/baseEvaluacion/base.html', context)

def agregarAlumno(request):
    print(request.POST)
    niveles = Nivel.objects.filter(estado="1")
    plan = PlanMedicionCurso.objects.get(pk=request.POST["plan"])
    listaIndicador = plan.indicador.all()
    alumnosEncontrados = RespuestaEvaluacion.objects.filter(codigoAlumno=request.POST["codigoAlumno"],
                                                     horario_id=request.POST["horario"],planMedicion_id=request.POST["plan"])
    if(len(alumnosEncontrados)>0):
        nombreActual= alumnosEncontrados[0].nombreAlumno
        return JsonResponse({"Insertado": True, "nombreActual":nombreActual},status=200)

    for indicador in listaIndicador:
        print(indicador.pk)
        nuevoAlumno = RespuestaEvaluacion.objects.create(nombreAlumno=request.POST["nombreAlumno"],
                                                     codigoAlumno=request.POST["codigoAlumno"],
                                                     horario_id=request.POST["horario"],
                                                     indicador_id=indicador.pk,
                                                     planMedicion_id=request.POST["plan"])
    ser_instance = serializers.serialize('json', [nuevoAlumno,])
    ser_instance2 = serializers.serialize('json', list(niveles), fields=('id', 'nombre', 'valor', 'estado'))
    return JsonResponse({}, status=200)

def guardarPuntuacion(request):
    alumno = RespuestaEvaluacion.objects.get(pk=request.POST["idAlumno"])
    descripcion = Rubrica.objects.get(indicador_id=request.POST["indicadorSeleccionado"],nivel_id=request.POST["nivelSeleccionado"]).descripcion
    if (alumno.valorNota == int(request.POST["nivelSeleccionado"])):
        alumno.descripcionP = ''
        alumno.valorNota = None
        alumno.rubrica_id = None
        alumno.calificado = 0
        alumno.save()
        return JsonResponse({}, status=200)
    alumno.descripcionP = descripcion
    alumno.valorNota = Nivel.objects.get(pk=request.POST["nivelSeleccionado"]).valor
    alumno.rubrica_id = Rubrica.objects.get(indicador_id=request.POST["indicadorSeleccionado"],nivel_id=request.POST["nivelSeleccionado"]).pk
    alumno.calificado = 1
    alumno.save()
    return JsonResponse({},status=200)

def editarAlumno(request):

    alumno = RespuestaEvaluacion.objects.get(pk=request.POST["idAlumno"])

    listaA= RespuestaEvaluacion.objects.filter(nombreAlumno=alumno.nombreAlumno,estado=1)
    nuevoCodigo = request.POST["codigoAlumno"]
    nuevoNombre = request.POST["nombreAlumno"]
    for als in listaA:
        als.codigoAlumno = nuevoCodigo
        als.nombreAlumno = nuevoNombre
        als.save()
    return JsonResponse({}, status=200)

def editarAlumnoNuevo(request):
    listaA= RespuestaEvaluacion.objects.filter(codigoAlumno= request.POST["codigoAlumno"],
                                               planMedicion_id=request.POST["plan"],
                                               horario_id=request.POST["horario"],
                                               estado=1)
    listaA.update(nombreAlumno = request.POST["nombreAlumno"])
    return JsonResponse({}, status=200)

def eliminarAlumno(request):
    horariopk= request.POST['horario']
    print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')
    print(horariopk)
    codigo_alumno = RespuestaEvaluacion.objects.filter(pk = request.POST["idAlumno"], horario_id=horariopk)[0].codigoAlumno
    alumnos = RespuestaEvaluacion.objects.filter(codigoAlumno= codigo_alumno, horario_id=horariopk)
    alumnos.update(estado = 0)
    return JsonResponse({}, status=200)

def muestraRubrica(request):
    niveles = Nivel.objects.filter(estado="1")
    rubrica = Rubrica.objects.filter(indicador_id=request.POST["indicadorSeleccionado"])
    indicador = Indicador.objects.get(pk=request.POST["indicadorSeleccionado"])
    ser_instance = serializers.serialize('json', list(rubrica),fields=('id','descripcion','especialidad','indicador','nivel'))
    ser_instance2 = serializers.serialize('json',[indicador,])
    ser_instance3 = serializers.serialize('json', list(niveles),fields=('id','nombre','valor','estado'))
    return JsonResponse({"rubrica": ser_instance,"indicador":ser_instance2, "niveles": ser_instance3 }, status=200)

def listarAlumno(request):
    filtrado = request.POST["filtrado"]
    especialidad = Especialidad.objects.get(pk=request.POST["especialidadSeleccionada"])
    planMedicion = PlanMedicion.objects.get(pk=request.POST["planMedicion"])
    niveles = Nivel.objects.filter(plaResultado_id = planMedicion.planResultados_id, estado="1",especialidad_id=especialidad.pk)
    if (filtrado!=""):
        if (filtrado.isnumeric()):
            listaAlumno = reversed(RespuestaEvaluacion.objects.filter(codigoAlumno__contains=filtrado, horario_id = request.POST["horarioSeleccionado"],indicador_id=request.POST["indicadorSeleccionado"], estado=1))
        else:
            listaAlumno = reversed(RespuestaEvaluacion.objects.filter(nombreAlumno__contains=filtrado, horario_id=request.POST["horarioSeleccionado"],indicador_id=request.POST["indicadorSeleccionado"], estado=1))
    else:
        listaAlumno = reversed(RespuestaEvaluacion.objects.filter(horario_id=request.POST["horarioSeleccionado"],indicador_id=request.POST["indicadorSeleccionado"], estado=1))

    ser_instance = serializers.serialize('json', list(listaAlumno),fields=('id', 'nombreAlumno', 'codigoAlumno', 'horario','calificado', 'valorNota','evidencia','archivo'))
    ser_instance2 = serializers.serialize('json', list(niveles), fields=('id', 'nombre', 'valor', 'estado'))
    return JsonResponse({"listaAlumno": ser_instance, "niveles": ser_instance2},  status=200)

def importarAlumno(request):
    # Funciona si guardas de nuevo el archivo descargado mediante campus porque dice que en este el formato y la
    # extension no coinciden, si se vuelve a guardar el archivo como xls si se puede guardar
    # wb = xlrd.open_workbook(filename=None, file_contents=request.FILES['archivo'].read())
    # table = wb.sheets()[0]
    # row = table.nrows
    # for i in range(7, row):
    #     col = table.row_values(i)
    #     if (col[0] == ''):
    #         break
    #     else:
    #         RespuestaEvaluacion.objects.create(nombreAlumno=col[1], codigoAlumno=int(col[0]),
    #                                                horario_id=request.POST["horariopk"])
    plan = PlanMedicionCurso.objects.get(pk=request.POST["plan"])
    listaIndicador = plan.indicador.all()
    archivo = request.FILES['archivo']
    data = pd.read_csv(archivo, sep="\t",skiprows=6,engine='c',encoding_errors='ignore',encoding='latin-1')
    for index,row in data.iterrows():

        alumnosEncontrados = RespuestaEvaluacion.objects.filter(codigoAlumno=row[0],
                                                                horario_id=request.POST["horariopk"],
                                                                planMedicion_id=request.POST["plan"])

        if (len(alumnosEncontrados) == 0):
            for indicador in listaIndicador:
                RespuestaEvaluacion.objects.create(nombreAlumno=row[1],
                                                               codigoAlumno=row[0],
                                                               horario_id=request.POST["horariopk"],
                                                               indicador_id=indicador.pk,
                                                               planMedicion_id=request.POST["plan"])
    return JsonResponse({}, status=200)

def subirEvidencia(request):
    # TODO::: Subir archivo a BBDD del PK del Evaluado
    print(request.POST["evaluadopk"])
    alumno = RespuestaEvaluacion.objects.get(pk=request.POST["evaluadopk"])
    listaA= RespuestaEvaluacion.objects.filter(codigoAlumno=alumno.codigoAlumno,nombreAlumno=alumno.nombreAlumno,estado=1)

    for als in listaA:
        als.evidencia = 1
        als.archivo = request.FILES["archivo"]
        als.save()
    # TODO::: Agregar una columna para guardar la evidencia una respuesta adecuada al front (Se hace con un IF y en dos JSONRESPONSE
    print(request.FILES["archivo"])
    return JsonResponse({}, status=200)

def eliminarEvidencia(request):
    print(request.POST)
    pk = request.POST['evidenciapk']
    alumno = RespuestaEvaluacion.objects.get(pk=request.POST["evidenciapk"])
    listaA = RespuestaEvaluacion.objects.filter(codigoAlumno=alumno.codigoAlumno, nombreAlumno=alumno.nombreAlumno,
                                                estado=1)
    for als in listaA:
        als.evidencia = 0
        os.remove(als.archivo.path)
        als.archivo = ""
        als.save()
    return JsonResponse({}, status=200)


def exportarMedicion(request):

    #Apertura de template.xlsx
    filename = os.path.join(settings.BASE_DIR, 'gestionarEvaluacion', 'Resource', 'template.xlsx')
    wb = load_workbook(filename)
    ws= wb.active

    horario = Horario.objects.get(pk=request.POST['cboHorario'])
    try:
        plan = PlanMedicionCurso.objects.get(pk=horario.curso.pk)
        plMedicion = PlanMedicion.objects.get(pk=plan.planMedicion_id)
        plResultados = PlanResultados.objects.get(pk=plMedicion.planResultados_id)
        indicadores_plan = PlanMedicionCurso.indicador.through.objects.filter(planmedicioncurso_id=plan.pk)
        indicadores = []
        for indicador in indicadores_plan:
            indicadores.append(Indicador.objects.get(pk=indicador.indicador_id))
        colIndicadores = []
        for i in range(len(indicadores)):
            if i==0:
                colIndicadores.append('E')
            else:
                colIndicadores.append(chr(ord(colIndicadores[i-1])+1))
        curso = Curso.objects.get(pk=plan.curso_id)
        semestre = Semestre.objects.get(pk=plan.semestre_id)

        #CABECERAS
        ws['E3'] = f'MEDICIÓN DEL CURSO -  {curso.nombre.upper()}'
        ws['D5'] = curso.nombre.upper()
        ws['G5'] = horario.codigo
        ws['D7'] = User.objects.get(id = horario.responsable.pk).first_name
        ws['K5'] = semestre.nombreCodigo


        #IMPRESORAS DE PUNTAJES

        alumnos = RespuestaEvaluacion.objects.filter(horario_id=request.POST['cboHorario'],estado='1').values('codigoAlumno').annotate(total=Count('codigoAlumno'))

        i=0
        for alumno in alumnos:
            ws[f'B{25 + i}'] = i + 1
            ws[f'C{25 + i}'] = alumno['codigoAlumno']
            nombreAlumno = RespuestaEvaluacion.objects.filter(horario_id=request.POST['cboHorario'],
                                                              estado='1',
                                                              codigoAlumno=alumno['codigoAlumno']
                                                             )[0].nombreAlumno
            ws[f'D{25 + i}'] = nombreAlumno

            celdaEstilo = ws[f'B{25 + i}']
            celdaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            celdaEstilo.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            celdaEstilo = ws[f'C{25 + i}']
            celdaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            celdaEstilo = ws[f'D{25 + i}']
            celdaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))

            i+=1

        r = lambda: random.randint(0, 255)
        colorRandom="%02X%02X%02X" % (r(),r(),r())

        while colorRandom<"71c46c":
            r = lambda: random.randint(0, 255)
            colorRandom = "%02X%02X%02X" % (r(), r(), r())

        for i in range(len(indicadores)):
            #CABECERA DE INDICADOR
            ws[f'{colIndicadores[i]}15'] = ws[f'{colIndicadores[i]}24'] = f'{indicadores[i].codigo}\n{indicadores[i].descripcion}'
            indicadorEstilo = ws[f'{colIndicadores[i]}15']
            indicadorEstilo.font = Font(name='Calibri', size=11, bold="True")
            indicadorEstilo.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            indicadorEstilo.fill=PatternFill(start_color=colorRandom, end_color=colorRandom, fill_type="solid")
            indicadorEstilo.border=Border(right=Side(border_style='medium'),left=Side(border_style='medium'),top=Side(border_style='medium'), bottom=Side(border_style='medium'))

            indicadorEstilo = ws[f'{colIndicadores[i]}24']
            indicadorEstilo.font = Font(name='Calibri', size=11, bold="True")
            indicadorEstilo.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            indicadorEstilo.fill = PatternFill(start_color=colorRandom, end_color=colorRandom, fill_type="solid")
            indicadorEstilo.border = Border(right=Side(border_style='medium'), left=Side(border_style='medium'),
                                            top=Side(border_style='medium'), bottom=Side(border_style='medium'))
            suma = 0
            n=0
            j=0
            for alumno in alumnos:
                valorNota = RespuestaEvaluacion.objects.filter(horario_id=request.POST['cboHorario'],
                                                           estado='1',
                                                           indicador_id= indicadores[i].pk,
                                                           codigoAlumno=alumno['codigoAlumno'])[0].valorNota
                if(valorNota != None):
                    suma = suma + valorNota
                    n= n+1
                else:
                    valorNota = 0
                ws[f'{colIndicadores[i]}{25 + j}'] = valorNota

                notaEstilo = ws[f'{colIndicadores[i]}{25 + j}']
                notaEstilo.font = Font(name='Calibri', size=11, bold=False)
                notaEstilo.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                notaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                j=j+1
            #Resultados
            niveles = Nivel.objects.filter(plaResultado_id=plResultados.pk, estado='1').order_by('-valor')
            valorMax = niveles[0].valor
            ws[f'{colIndicadores[i]}18'] = n
            ws[f'{colIndicadores[i]}16'] = suma/n if n>0 else 0
            ws[f'{colIndicadores[i]}17'] = suma/n/valorMax if n>0 else 0

            celdaEstilo=ws[f'{colIndicadores[i]}18']
            celdaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

            celdaEstilo = ws[f'{colIndicadores[i]}16']
            celdaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            celdaEstilo.number_format = '0.00'

            celdaEstilo = ws[f'{colIndicadores[i]}17']
            celdaEstilo.border = Border(right=Side(border_style='thin'), left=Side(border_style='thin'),
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            celdaEstilo.number_format = '0.00%'




    #=======================================NO BORRAR - NO BORRAR====================================
        # ws.merge_cells('O15:Q15')
        # ws['O15'].font = Font(bold="True")
        # ws['O15'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        # ws['O15'] = "CAGASTES"
        # currentCell = ws['O15']
        # currentCell.alignment = Alignment(horizontal='center',vertical='center')
        # currentCell.border = Border(left=Side(border_style='medium'),
        #                             right=Side(border_style='medium'),
        #                             top=Side(border_style='medium'),
        #                             bottom=Side(border_style='medium'))
        # ws.row_dimensions[16].height = 70
        # ws.column_dimensions['S'].width = 70
        #thin : borde simple
        # =======================================NO BORRAR - NO BORRAR====================================

        # #Establecer el nombre del archivo
    except:
        print("ERROR en generacion de reporte")

    nombre_archivo = f'Medicion-{horario.codigo}.xlsx'

    #Definir el tipo de respuesta que se va a dar
    response= HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def historial(request):
    context = {
        'listaSemestre': Semestre.objects.all()
    }
    return render(request, 'gestionarEvaluacion/historialEvaluacion.html',context)

def listarCursoMedicion(request):
    ##Estado 2 de Terminado
    ##Cambiar especialidad_id por el que deberia o si se pondra comboBox mejor :D
    planMedicion = PlanMedicion.objects.filter(estado='2',especialidad_id=5).filter(semestre=request.POST['semestre'])
    e = list(planMedicion)
    for i in range(len(e)):
        print(e[i].nombre)
    print(planMedicion)
    try:
        listaCursos = []
        planMedicionCursos = PlanMedicionCurso.objects.filter(semestre=request.POST['semestre'], planMedicion=planMedicion[0].pk)
        for planMedicionCurso in planMedicionCursos:
            curso = Curso.objects.get(pk=planMedicionCurso.curso_id, especialidad_id=5)
            print(curso.nombre)
            listaCursos.append(curso)
            # if curso is not None:
            #    listaCursos.append(curso)
    except:
        print("No existe plan de medicion asociado!!!")
    print("#########")
    print("Cantidad de cursos:")
    print(len(listaCursos))
    for i in range(len(listaCursos)):
        print(listaCursos[i].nombre)
    print("#########")
    ser_instance = serializers.serialize('json', listaCursos)
    print(ser_instance)
    return JsonResponse({"cursoLista": ser_instance}, status=200)

def algoritmoSacarPorcentajes():
    esp = 5
    ciclo = "2020-2"
    # objeto semetre, niveles y especialidad
    semAux = Semestre.objects.get(nombreCodigo=ciclo)
    niveles = Nivel.objects.filter(especialidad_id=esp, estado='1').order_by('-valor')
    planSemestre = PlanMedicion.semestre.through.objects.filter(semestre_id=semAux.pk)

    # inicializando data
    valorMax = niveles[0].valor
    planCal = None

    for p in planSemestre:
        estado = [1, 2]
        planesMedi = PlanMedicion.objects.filter(pk=p.planmedicion_id, estado__in=estado, especialidad_id=esp)
        if planesMedi:
            planCal = planesMedi[0]
            break

    if planCal:
        planRes = PlanResultados.objects.get(pk=planCal.planResultados.pk)
        resultados = ResultadoPUCP.objects.filter(planResultado_id=planRes.pk, estado=1)
        for r in resultados:
            print(r.codigo, r.descripcion)
            indicadores = Indicador.objects.filter(resultado_id=r.pk, estado=1)
            for i in indicadores:
                print(i.codigo, i.descripcion)
                plMedCur = PlanMedicionCurso.objects.filter(planMedicion_id=planCal.pk, semestre_id=semAux.pk, estado=1)
                n = 0
                for pl in plMedCur:
                    planMedCurInt = PlanMedicionCurso.indicador.through.objects.filter(indicador_id=i.pk,
                                                                                       planmedicioncurso_id=pl.pk)
                    if planMedCurInt:
                        n += 1
                        horarios = Horario.objects.filter(curso_id=pl.pk, estado=1)
                        cantAlum = 0
                        acumNota = 0
                        for h in horarios:
                            alumnosEva = RespuestaEvaluacion.objects.filter(horario_id=h.pk,
                                                                            estado='1',
                                                                            indicador_id=i.pk)
                            for alum in alumnosEva:
                                valorNota = alum.valorNota
                                if valorNota:
                                    cantAlum += 1
                                    acumNota += valorNota

                        cursoAsig = Curso.objects.get(pk=pl.curso.pk)
                        valorPorcentaje = acumNota / cantAlum / valorMax if cantAlum > 0 else 0
                        print(cursoAsig.nombre, valorPorcentaje)
    else:
        print("No salio nada")
        print("==============================")