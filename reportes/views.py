from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import JsonResponse
from django.core import serializers
# Create your views here.
from gestionarCurso.models import Curso
from gestionarEspecialidad.models import Asistente, Auditor, Especialidad
from django.contrib.auth.models import Group

from gestionarFacultad.models import Facultad
from gestionarSemestre.models import Semestre
import os
from openpyxl import Workbook,load_workbook,styles
from django.http.response import HttpResponse
from django.shortcuts import render, redirect
from django.http import JsonResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from validate_email import validate_email
from django.urls import reverse
from django.views.generic import FormView
from authentication.forms import ChangePasswordForm
from django.http import HttpResponseRedirect
from authentication.models import User
from demosoftware3.settings import MEDIA_URL
from gestionarFacultad.views import listarFacultad
from django.contrib.auth import authenticate
from django.contrib.auth import login, logout
from django.contrib.auth.models import Group
from demosoftware3 import settings
import uuid
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.template.loader import render_to_string
import pandas as pd
import math

from gestionarIndicadores.models import Indicador
from gestionarPlanMedicion.models import PlanMedicion, PlanMedicionCurso
from gestionarResultados.models import ResultadoPUCP, PlanResultados

from gestionarSemestre.models import Semestre

from gestionarHorario.models import Horario

from gestionarEvaluacion.models import RespuestaEvaluacion

from gestionarNiveles.models import Nivel



from gestionarEspecialidad.models import Especialidad

from gestionarRubrica.models import Rubrica



@login_required
def reportes(request):
    especialidades = []
    facultades = []
    semestres = Semestre.objects.filter()

    if (request.user.rol_actual == "Asistente de acreditación"):
        usuario = request.user
        especialidades = []
        items = Asistente.objects.filter(user_id=usuario.pk)
        for item in items:
            especialidades.append(item.especialidad)
        context = {
            'semestres':semestres,
            'especialidades': especialidades,
        }
        return render(request, 'reportes/reportesCE.html', context)

    if (request.user.rol_actual == "Auditor"):
        usuario = request.user
        especialidades = []
        items = Auditor.objects.filter(user_id=usuario.pk)
        for item in items:
            especialidades.append(item.especialidad)
        context = {
            'semestres': semestres,
            'especialidades': especialidades,
        }
        return render(request, 'reportes/reportesCE.html', context)

    if (request.user.rol_actual == "Coordinador de facultad"):
        usuario = request.user
        grupo = Group.objects.get(pk=4)
        facultades = Facultad.objects.filter(responsable=usuario.pk)
        context = {
            'semestres': semestres,
            'facultades': facultades,
        }
        return render(request, 'reportes/reportesCF.html', context)

    if (request.user.rol_actual == "Coordinador de especialidad"):
        usuario = request.user
        grupo = Group.objects.get(pk=5)
        especialidades = Especialidad.objects.filter(responsable=usuario.pk)
        context = {
            'semestres': semestres,
            'especialidades': especialidades,
        }
        return render(request, 'reportes/reportesCE.html', context)

def obtenerEspecialidades(request):
    id_facultad = request.POST['facultad']
    especialidades = Especialidad.objects.filter(facultad_id=id_facultad, estado='1')
    data = serializers.serialize("json", especialidades)
    return JsonResponse({"resp": data}, status=200)


def obtenerNiveles(esp,ciclo):
    semAux = Semestre.objects.get(pk=ciclo)
    planSemestre = PlanMedicion.semestre.through.objects.filter(semestre_id=semAux.pk)
    planCal = None
    for p in planSemestre:
        estado = [1, 2]
        planesMedi = PlanMedicion.objects.filter(pk=p.planmedicion_id, estado__in=estado, especialidad_id=esp)
        if planesMedi:
            planCal = planesMedi[0]
            break

    plResultado = PlanResultados.objects.get(pk=planCal.planResultados.pk)
    return plResultado


def generarReportes(request):
    if request.POST:

        esp = request.POST['cboEspecialidad']
        ciclo = request.POST['cboSemestre']
        print("esp", esp)
        print("ciclo", ciclo)

        if '_rubricas' in request.POST:
            print("rubricas")
            semestre = Semestre.objects.get(pk=ciclo)
            #Apertura de template.xlsx
            filename = os.path.join(settings.BASE_DIR, 'reportes', 'Resources', 'template.xlsx')
            wb = load_workbook(filename)
            ws = wb.active
            try:
                nivelesX = []
                nivelesY=[]
                especialidad = Especialidad.objects.get(pk=esp)
                plan = obtenerNiveles(esp,ciclo)
                niveles = Nivel.objects.filter(plaResultado_id=plan.pk,estado='1').order_by('-valor')
                for n in range(len(niveles)):
                    if n == 0:
                        nivelesX.append('D')
                        nivelesY.append('B')
                    else:
                        nivelesX.append(chr(ord(nivelesX[n - 1]) + 1))
                    ws[f'{nivelesX[n]}7'] = f'{niveles[n].nombre} ({niveles[n].valor})'
                    currentcell = ws[f'{nivelesX[n]}7']
                    currentcell.font = Font(name='Calibri', size=11, bold="True")
                    currentcell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    currentcell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                    if n == 0:
                        currentcell.border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'),
                                                    bottom=Side(border_style='thin'))
                    elif n == len(niveles) - 1:
                        currentcell.border = Border(right=Side(border_style='thin'), top=Side(border_style='thin'),
                                                    bottom=Side(border_style='thin'))
                    else:
                        currentcell.border = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                ws.merge_cells(f'D6:{nivelesX[len(niveles) - 1]}6')
                ws.merge_cells(f'B5:{nivelesX[len(niveles) - 1]}5')

                ws['C3'] = f'RÚBRICAS DEL PROGRAMA DE {especialidad.nombre.upper()}'

                semAux = Semestre.objects.get(pk=ciclo)
                planSemestre = PlanMedicion.semestre.through.objects.filter(semestre_id=semAux.pk)
                planCal = None
                for p in planSemestre:
                    estado = [1, 2]
                    planesMedi = PlanMedicion.objects.filter(pk=p.planmedicion_id, estado__in=estado, especialidad_id=esp)
                    if planesMedi:
                        planCal = planesMedi[0]
                        break

                plResultado = PlanResultados.objects.filter(pk=planCal.planResultados.pk)
                if plResultado:
                    for plR in plResultado:
                        resTodos = ResultadoPUCP.objects.filter(planResultado_id=plR.pk, estado=1)
                        for i in range(len(resTodos)):
                            res = resTodos[i]
                            codResul = res.codigo
                            ws.title = codResul
                            if i < len(resTodos) - 1:
                                wb.copy_worksheet(ws)
                                nombreAux = codResul + " Copy"
                                ws = wb[nombreAux]

                            wsAux = wb[codResul]
                            wsAux['B5'] = f'{res.codigo} ({res.descripcion})'
                            wsAux['B5'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                            filaInicio = 8
                            indImpr = Indicador.objects.filter(resultado_id=resTodos[i].pk, estado=1)
                            if indImpr:
                                for numInd in range(len(indImpr)):
                                    wsAux[f'C{filaInicio}'] = f'{indImpr[numInd].codigo} {indImpr[numInd].descripcion}'
                                    celdaEstilo = wsAux[f'C{filaInicio}']
                                    celdaEstilo.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
                                    celdaEstilo.border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'),
                                                                bottom=Side(border_style='thin'))
                                    celdaEstilo.font = Font(name='Calibri', size=12, bold="True")
                                    for n in range(len(niveles)):
                                        rubrica = Rubrica.objects.filter(indicador_id=indImpr[numInd].pk,
                                                                         nivel_id=niveles[n].pk)
                                        if rubrica:
                                            wsAux[f'{nivelesX[n]}{filaInicio}'] = rubrica[0].descripcion
                                        rubricaEstilo = wsAux[f'{nivelesX[n]}{filaInicio}']
                                        rubricaEstilo.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
                                        rubricaEstilo.font = Font(name='Calibri', size=11, bold=False)
                                        if n == len(niveles) - 1:
                                            rubricaEstilo.border = Border(left=None,
                                                                          right=Side(border_style='thin'),
                                                                          top=Side(border_style='thin'),
                                                                          bottom=Side(border_style='thin'))
                                        else:
                                            rubricaEstilo.border = Border(left=None,
                                                                          right=None,
                                                                          top=Side(border_style='thin'),
                                                                          bottom=Side(border_style='thin'))
                                    filaInicio += 1
                            if filaInicio > 8:
                                wsAux.merge_cells(f'B8:B{filaInicio - 1}')
            except:
                print("El ciclo no está siendo medido ")

            nom_arch = f'Rubrica {especialidad.nombre}-{semestre.nombreCodigo}.xlsx'
            # Definir el tipo de respuesta que se va a dar
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename = {0}".format(nom_arch)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response
        elif '_resultados' in request.POST:
            print("resultados")
            filename = os.path.join(settings.BASE_DIR, 'reportes', 'Resources', 'template2.xlsx')
            wb = load_workbook(filename)
            ws = wb.active
            semAux = Semestre.objects.get(pk=ciclo)
            especialidad = Especialidad.objects.get(pk=esp)
            try:
                plan = obtenerNiveles(esp,ciclo)
                niveles = Nivel.objects.filter(plaResultado_id=plan.pk,estado='1').order_by('-valor')
                valorMax = niveles[0].valor
                planSemestre = PlanMedicion.semestre.through.objects.filter(semestre_id=semAux.pk)
                planCal = None
                for p in planSemestre:
                    estado = [1, 2]
                    planesMedi = PlanMedicion.objects.filter(pk=p.planmedicion_id, estado__in=estado, especialidad_id=esp)
                    if planesMedi:
                        planCal = planesMedi[0]
                        break

                if planCal:
                    planRes = PlanResultados.objects.get(pk=planCal.planResultados.pk)
                    resultados = ResultadoPUCP.objects.filter(planResultado_id=planRes.pk, estado=1)
                    filaInicio = 8
                    for r in resultados:
                        print(r.codigo, r.descripcion)
                        indicadores = Indicador.objects.filter(resultado_id=r.pk, estado=1)
                        totalCursosValidos=0
                        for i in indicadores:
                            print(i.codigo, i.descripcion)
                            plMedCur = PlanMedicionCurso.objects.filter(planMedicion_id=planCal.pk, semestre_id=semAux.pk,
                                                                        estado=1)
                            cursosValidos = 0
                            n = 0
                            for pl in plMedCur:
                                planMedCurInt = PlanMedicionCurso.indicador.through.objects.filter(indicador_id=i.pk,
                                                                                                   planmedicioncurso_id=pl.pk)
                                if planMedCurInt:
                                    n += 1
                                    horarios = Horario.objects.filter(curso_id=pl.pk, estado=1)
                                    cantAlum = 0
                                    acumNota = 0
                                    for h in horarios:
                                        alumnosEva = RespuestaEvaluacion.objects.filter(horario_id=h.pk,
                                                                                        estado='1',
                                                                                        indicador_id=i.pk)
                                        for alum in alumnosEva:
                                            valorNota = alum.valorNota
                                            if valorNota:
                                                cantAlum += 1
                                                acumNota += valorNota

                                    print(pl.curso_id)
                                    cursoAsig = Curso.objects.get(pk=pl.curso_id)
                                    valorPorcentaje = acumNota / cantAlum / valorMax if cantAlum > 0 else 0
                                    print(cursoAsig.nombre, valorPorcentaje)
                                    if horarios:
                                        cursosValidos += 1
                                        ws[f'D{filaInicio+totalCursosValidos+cursosValidos-1}']=cursoAsig.nombre
                                        ws[f'E{filaInicio+totalCursosValidos+cursosValidos-1}']=valorPorcentaje


                            if cursosValidos>0:
                                ws.merge_cells(f'C{filaInicio+totalCursosValidos}:C{filaInicio+totalCursosValidos+cursosValidos-1}')
                                ws[f'C{filaInicio+totalCursosValidos}'] = f'({i.codigo}) {i.descripcion}'
                                totalCursosValidos+=cursosValidos

                        if totalCursosValidos>0:
                            ws.merge_cells(f'B{filaInicio}:B{filaInicio + totalCursosValidos - 1}')
                            ws[f'B{filaInicio}']=r.codigo
                            filaInicio += totalCursosValidos
                else:
                    print("No existe plan de medición asociado al ciclo")
            except:
                print("Hubo un error")
            nom_arch = f'Reporte {especialidad.nombre}-{semAux.nombreCodigo}.xlsx'
            # Definir el tipo de respuesta que se va a dar
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename = {0}".format(nom_arch)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response